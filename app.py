import os
import sys
import pandas as pd
import numpy as np
import datetime
from flask import Flask, render_template, request, send_file, send_from_directory
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import traceback

app = Flask(__name__)

# Determine if running as a script or frozen (PyInstaller)
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
    # When frozen, templates and static files are in sys._MEIPASS
    app = Flask(__name__, template_folder=os.path.join(BASE_DIR, 'templates'))
else:
    BASE_DIR = os.path.abspath(os.path.dirname(__file__))
    app = Flask(__name__)

UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads') # Use CWD for user-accessible folders
DOWNLOAD_FOLDER = os.path.join(os.getcwd(), 'downloads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['DOWNLOAD_FOLDER'] = DOWNLOAD_FOLDER

class ExecutionLogger:
    def __init__(self):
        self.logs = []
    
    def user(self, message, level='INFO'):
        self._add_log(level, message, 'user')
        
    def dev(self, message, level='DEBUG'):
        self._add_log(level, message, 'dev')
        
    def _add_log(self, level, message, category):
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        self.logs.append({
            'time': timestamp, 
            'level': level, 
            'msg': message,
            'category': category
        })

def parse_server_pairs(text):
    pairs = []
    seen = set()
    duplicates = []
    
    if not text:
        return pairs, duplicates
        
    lines = text.strip().split('\n')
    for line in lines:
        parts = line.replace('，', ',').split(',')
        if len(parts) >= 2:
            try:
                s1 = int(parts[0].strip())
                s2 = int(parts[1].strip())
                
                # Sort tuple to treat (A, B) same as (B, A)
                pair_key = tuple(sorted((s1, s2)))
                
                if pair_key in seen:
                    duplicates.append(f"{s1} ↔ {s2}")
                else:
                    seen.add(pair_key)
                    pairs.append((s1, s2))
            except ValueError:
                continue
    return pairs, duplicates

def get_server_info(df, server_id):
    row = df[df['区服ID'] == server_id]
    if row.empty:
        return None
    return row.iloc[0]

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        logger = ExecutionLogger()
        try:
            logger.user("开始处理任务...")
            logger.dev("初始化请求参数解析")
            
            # 1. Save files
            csv_files = request.files.getlist('csv_files')
            xlsx_file = request.files['xlsx_file']
            pairs_text = request.form['pairs_text']

            if not csv_files or not xlsx_file:
                return "Missing files", 400

            xlsx_path = os.path.join(app.config['UPLOAD_FOLDER'], 'input.xlsx')
            xlsx_file.save(xlsx_path)
            logger.user("合服计划表 (XLSX) 上传成功")

            # 2. Process CSVs (Merge Multiple)
            logger.user(f"正在处理 {len(csv_files)} 个服务器数据文件...")
            dfs = []
            for i, file in enumerate(csv_files):
                if file.filename == '':
                    continue
                temp_path = os.path.join(app.config['UPLOAD_FOLDER'], f'input_{i}.csv')
                file.save(temp_path)
                try:
                    # 尝试读取 CSV，跳过第一行
                    df_temp = pd.read_csv(temp_path, header=1)
                    dfs.append(df_temp)
                    logger.dev(f"读取 CSV {file.filename} 成功，行数: {len(df_temp)}")
                except Exception as e:
                    logger.user(f"读取文件 {file.filename} 失败", 'ERROR')
                    logger.dev(f"CSV 读取异常: {str(e)}", 'ERROR')
            
            if not dfs:
                 return "没有有效的 CSV 文件", 400
                 
            df = pd.concat(dfs, ignore_index=True)
            logger.user(f"数据合并完成，共 {len(df)} 条记录")
            
            # Ensure numeric columns
            cols_to_numeric = ['区服ID', '前2名战力之和', '最高玩家累充金额', 'DAU', '跨服ID', 'code', '有效DAU', '当天付费账号数', '峰值在线', 'MAC_DAU', 'IP_DAU', '账号DAU', '总注册角色']
            for col in cols_to_numeric:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                    if col in ['区服ID', 'DAU', '跨服ID', 'code', '总注册角色', '峰值在线', '当天付费账号数']: # Explicitly cast ID-like or count-like fields to int
                         try:
                            df[col] = df[col].astype(int)
                         except:
                            pass # Keep as float if int conversion fails (e.g. too large or weird values)

            # Sort
            logger.dev("执行数据排序: 前2名战力之和 (降序)")
            df = df.sort_values(by='前2名战力之和', ascending=False).reset_index(drop=True)
            df['真实排名'] = df.index + 1
            total_servers = len(df)
            
            input_pairs, duplicates = parse_server_pairs(pairs_text)
            
            if duplicates:
                logger.user(f"发现并忽略 {len(duplicates)} 组重复检测对", 'WARN')
                if len(duplicates) <= 5:
                    for dup in duplicates:
                        logger.dev(f"忽略重复: {dup}", 'WARN')
                else:
                    logger.dev(f"重复列表 (前5个): {', '.join(duplicates[:5])}...", 'WARN')
            
            logger.user(f"解析输入：共 {len(input_pairs)} 组有效检测区服")
            
            alert_groups = [] 
            normal_groups = [] 
            
            # 3. Primary Alert Check
            logger.dev("开始执行初级警报检测 (Primary Check)")
            for s1, s2 in input_pairs:
                row1 = get_server_info(df, s1)
                row2 = get_server_info(df, s2)
                
                if row1 is None:
                    logger.user(f"警告：区服 {s1} 数据缺失，已跳过", 'WARN')
                if row2 is None:
                    logger.user(f"警告：区服 {s2} 数据缺失，已跳过", 'WARN')

                if row1 is None or row2 is None:
                    continue
                    
                rank1 = row1['真实排名']
                rank2 = row2['真实排名']
                
                # Conditions
                cond_a = abs(rank1 - rank2) <= 5
                
                top_25_threshold = total_servers * 0.25
                cond_b = (rank1 <= top_25_threshold and rank2 <= top_25_threshold and
                          row1['最高玩家累充金额'] >= 5000 and row2['最高玩家累充金额'] >= 5000)
                          
                power1 = row1['前2名战力之和']
                power2 = row2['前2名战力之和']
                cond_c = abs(power1 - power2) <= 1000000000
                
                if cond_a or cond_b or cond_c:
                    reasons = []
                    if cond_a: reasons.append(f"排名接近(差{abs(rank1-rank2)})")
                    if cond_b: reasons.append("高战高充(前25%)")
                    if cond_c: reasons.append("战力接近(差<=10亿)")
                    reason_str = "; ".join(reasons)
                    
                    logger.user(f"发现警报：{s1} 和 {s2} - {reason_str}", 'WARN')
                    alert_groups.append({'ids': [s1, s2], 'reason': reason_str})
                else:
                    normal_groups.append((s1, s2))
            
            logger.user(f"检测完成：发现 {len(alert_groups)} 组警报，{len(normal_groups)} 组正常")

            # 4. Secondary Alert Check
            logger.dev("加载 XLSX 进行二次关联检测")
            wb = load_workbook(xlsx_path)
            ws = wb.active
            
            header_row = [cell.value for cell in ws[1]]
            try:
                target_col_idx = header_row.index('目标服') 
                part_col_idx = header_row.index('参与服')
            except ValueError:
                target_col_idx = 0
                part_col_idx = 1

            def find_partner_in_xlsx(server_id):
                for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    t_id = row[target_col_idx]
                    p_id = row[part_col_idx]
                    if t_id == server_id: return r_idx, p_id
                    if p_id == server_id: return r_idx, t_id
                return None, None

            final_alert_rows = [] 
            secondary_alert_groups = []
            
            for group in alert_groups:
                group_id = f"Group_{group['ids'][0]}_{group['ids'][1]}"
                for sid in group['ids']:
                    r = get_server_info(df, sid)
                    if r is not None:
                        r_dict = r.to_dict()
                        r_dict['警报组ID'] = group_id
                        r_dict['警报原因'] = group['reason']
                        
                        # Force int type for ID fields in dict if they became float
                        for k in ['区服ID', 'DAU', '跨服ID', 'code', '总注册角色', '峰值在线', '当天付费账号数']:
                            if k in r_dict:
                                try:
                                    if isinstance(r_dict[k], float):
                                        r_dict[k] = int(r_dict[k])
                                    elif isinstance(r_dict[k], (int, float)): # Ensure even numpy ints are python ints
                                         r_dict[k] = int(r_dict[k])
                                except:
                                    pass

                        final_alert_rows.append(r_dict)
                        
                s1, s2 = group['ids'][0], group['ids'][1]
                partners = []
                _, p1 = find_partner_in_xlsx(s1)
                if p1: partners.append(p1)
                _, p2 = find_partner_in_xlsx(s2)
                if p2: partners.append(p2)
                
                # Independent Secondary Checks
                # Logic: For each server in the alerted pair (S1, S2), check their respective partners.
                # If a Low DAU situation is found (either the server itself or its partner),
                # create a NEW, INDEPENDENT alert group for that pair (Server + Partner).
                
                # Helper to process secondary pair
                def process_secondary_pair(main_id, partner_id):
                    if not partner_id:
                        return

                    main_row = get_server_info(df, main_id)
                    partner_row = get_server_info(df, partner_id)
                    
                    # Check DAU conditions
                    alerts = []
                    if main_row is not None and main_row['DAU'] <= 5:
                        alerts.append(f"{main_id}本身DAU过低({int(main_row['DAU'])})")
                    if partner_row is not None and partner_row['DAU'] <= 5:
                        alerts.append(f"关联服{partner_id}DAU过低({int(partner_row['DAU'])})")
                        
                    if alerts:
                        # Create a unique group for this secondary relationship
                        # Sort IDs to ensure consistent Group ID (e.g. Group_Small_Big)
                        pair_ids = sorted([main_id, partner_id])
                        sec_group_id = f"Group_{pair_ids[0]}_{pair_ids[1]}"
                        reason_str = " | ".join(alerts)
                        
                        # Log it
                        logger.user(f"触发独立二次警报: {sec_group_id} - {reason_str}", 'WARN')
                        
                        # Add to summary list for frontend
                        secondary_alert_groups.append({
                            'ids': pair_ids,
                            'reason': reason_str
                        })
                        
                        # Add rows to CSV data
                        # 1. Add Main Server Row
                        if main_row is not None:
                            r_dict = main_row.to_dict()
                            r_dict['警报组ID'] = sec_group_id
                            r_dict['警报原因'] = reason_str
                            # Force int type for ID fields in dict if they became float
                            for k in ['区服ID', 'DAU', '跨服ID', 'code', '总注册角色', '峰值在线', '当天付费账号数']:
                                if k in r_dict and isinstance(r_dict[k], float):
                                    r_dict[k] = int(r_dict[k])
                            final_alert_rows.append(r_dict)
                            
                        # 2. Add Partner Row
                        if partner_row is not None:
                            pr_dict = partner_row.to_dict()
                            pr_dict['警报组ID'] = sec_group_id
                            pr_dict['警报原因'] = reason_str
                            # Force int type for ID fields in dict if they became float
                            for k in ['区服ID', 'DAU', '跨服ID', 'code', '总注册角色', '峰值在线', '当天付费账号数']:
                                if k in pr_dict and isinstance(pr_dict[k], float):
                                    pr_dict[k] = int(pr_dict[k])
                            final_alert_rows.append(pr_dict)

                # Check S1 and its partner
                _, p1 = find_partner_in_xlsx(s1)
                process_secondary_pair(s1, p1)

                # Check S2 and its partner
                _, p2 = find_partner_in_xlsx(s2)
                process_secondary_pair(s2, p2)

            # Create Alert CSV with optimized formatting
            if final_alert_rows:
                alert_df = pd.DataFrame(final_alert_rows)
                
                # Define columns to keep
                # Base cols from user requirement
                base_cols_to_keep = [
                    '区服ID', 'DAU', '近3日收入', '近7日收入', 
                    '第一名战力', '第二名战力', '第三名战力', 
                    '前2名战力之和', '前3名战力之和', 
                    '前十平均战力', '前十平均等级', '最高玩家累充金额'
                ]
                # Added cols by logic
                added_cols_to_keep = ['真实排名', '警报组ID', '警报原因']
                
                all_keep_cols = added_cols_to_keep + base_cols_to_keep
                
                # Filter columns: intersection of what we want and what exists
                final_cols = [c for c in all_keep_cols if c in alert_df.columns]
                
                alert_df = alert_df[final_cols]
                
                # Add empty rows between groups for visual separation
                # Convert to list of dicts to easily insert rows
                # Sort by Group ID to ensure they are contiguous, and then by Real Rank
                alert_df.sort_values(by=['警报组ID', '真实排名'], ascending=[True, True], inplace=True)
                
                output_rows = []
                current_group = None
                
                for _, row in alert_df.iterrows():
                    if current_group is not None and row['警报组ID'] != current_group:
                        # Insert empty row (dict with all empty strings to prevent float promotion)
                        empty_row = {c: "" for c in alert_df.columns}
                        output_rows.append(empty_row)
                    
                    output_rows.append(row.to_dict())
                    current_group = row['警报组ID']
                
                final_df = pd.DataFrame(output_rows)
                output_csv_path = os.path.join(app.config['DOWNLOAD_FOLDER'], 'alert_result.csv')
                final_df.to_csv(output_csv_path, index=False, encoding='utf-8-sig')
            else:
                pd.DataFrame().to_csv(os.path.join(app.config['DOWNLOAD_FOLDER'], 'alert_result.csv'), index=False)

            # 5. Merge Servers (Merge Requests)
            logger.user("正在处理正常组的合并申请...")
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            swapped_log_data = [] 

            server_row_map = {}
            for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                t_id = row[target_col_idx]
                p_id = row[part_col_idx]
                if t_id: server_row_map[t_id] = r_idx
                if p_id: server_row_map[p_id] = r_idx
                
            actual_swapped_count = 0
            
            for s1, s2 in normal_groups:
                r1_idx = server_row_map.get(s1)
                r2_idx = server_row_map.get(s2)
                
                if r1_idx and r2_idx and r1_idx != r2_idx:
                    actual_swapped_count += 1
                    
                    # Capture State Before Merge
                    c1_t = ws.cell(row=r1_idx, column=target_col_idx+1)
                    c1_p = ws.cell(row=r1_idx, column=part_col_idx+1)
                    v1_t, v1_p = c1_t.value, c1_p.value
                    
                    c2_t = ws.cell(row=r2_idx, column=target_col_idx+1)
                    c2_p = ws.cell(row=r2_idx, column=part_col_idx+1)
                    v2_t, v2_p = c2_t.value, c2_p.value
                    
                    # Helper to format server pair string
                    def fmt_pair(t, p):
                        return f"[{t if t else '空'} + {p if p else '空'}]"

                    before_str_1 = fmt_pair(v1_t, v1_p)
                    before_str_2 = fmt_pair(v2_t, v2_p)

                    # --- MERGE LOGIC ---
                    # Goal: Put s1 and s2 into Row 1. Put their leftovers (partners) into Row 2.
                    
                    # 1. Identify partners (leftovers)
                    # If v1_t is s1, then v1_p is the partner. And vice versa.
                    p1 = v1_p if v1_t == s1 else v1_t
                    p2 = v2_p if v2_t == s2 else v2_t
                    
                    # 2. Assign new pairs
                    # Row 1 gets s1 and s2 (The requested pair)
                    # Row 2 gets p1 and p2 (The leftover pair)
                    
                    # Sort pairs (Small ID first)
                    new_pair_1 = sorted([x for x in [s1, s2] if x is not None])
                    new_pair_2 = sorted([x for x in [p1, p2] if x is not None])
                    
                    # 3. Update Cells
                    # Row 1
                    if len(new_pair_1) == 2:
                        c1_t.value, c1_p.value = new_pair_1[0], new_pair_1[1]
                        final_v1_t, final_v1_p = new_pair_1[0], new_pair_1[1]
                    elif len(new_pair_1) == 1:
                         c1_t.value, c1_p.value = new_pair_1[0], None
                         final_v1_t, final_v1_p = new_pair_1[0], None
                    else:
                         c1_t.value, c1_p.value = None, None # Should not happen for s1, s2
                         final_v1_t, final_v1_p = None, None

                    # Row 2
                    if len(new_pair_2) == 2:
                        c2_t.value, c2_p.value = new_pair_2[0], new_pair_2[1]
                        final_v2_t, final_v2_p = new_pair_2[0], new_pair_2[1]
                    elif len(new_pair_2) == 1:
                         c2_t.value, c2_p.value = new_pair_2[0], None
                         final_v2_t, final_v2_p = new_pair_2[0], None
                    else:
                         c2_t.value, c2_p.value = None, None
                         final_v2_t, final_v2_p = None, None

                    # Capture State After Swap
                    after_str_1 = fmt_pair(final_v1_t, final_v1_p)
                    after_str_2 = fmt_pair(final_v2_t, final_v2_p)
                    
                    # Human readable change log
                    change_log = (
                        f"组1 (行{r1_idx}): {before_str_1} ➔ {after_str_1} (合并目标)\n"
                        f"   组2 (行{r2_idx}): {before_str_2} ➔ {after_str_2} (剩余自动组队)"
                    )
                    
                    logger.user(f"✅ 成功合并 {s1} + {s2}\n   {change_log}", 'SUCCESS')
                    
                    swapped_log_data.append({
                        '合并申请': f"{s1}+{s2}",
                        '原始行号1': r1_idx, '原始行号2': r2_idx,
                        'Before1': before_str_1, 'After1': after_str_1,
                        'Before2': before_str_2, 'After2': after_str_2,
                        '状态': '已合并'
                    })
                        
                    for cell in ws[r1_idx]: cell.fill = fill
                    for cell in ws[r2_idx]: cell.fill = fill

                    # Update Map
                    # Row 1 now contains s1 and s2
                    if s1: server_row_map[s1] = r1_idx
                    if s2: server_row_map[s2] = r1_idx
                    # Row 2 now contains p1 and p2
                    if p1: server_row_map[p1] = r2_idx
                    if p2: server_row_map[p2] = r2_idx
                    
                    logger.dev(f"执行合并 ({s1}, {s2}) -> Row {r1_idx}, Leftovers ({p1}, {p2}) -> Row {r2_idx}")
                else:
                    logger.dev(f"无法合并 ({s1}, {s2}): 未找到匹配行或已在同一行")

            if swapped_log_data:
                swapped_df = pd.DataFrame(swapped_log_data)
                output_swapped_path = os.path.join(app.config['DOWNLOAD_FOLDER'], 'swapped_log.csv')
                swapped_df.to_csv(output_swapped_path, index=False, encoding='utf-8-sig')
            else:
                pd.DataFrame().to_csv(os.path.join(app.config['DOWNLOAD_FOLDER'], 'swapped_log.csv'), index=False)

            output_xlsx_path = os.path.join(app.config['DOWNLOAD_FOLDER'], 'result_plan.xlsx')
            wb.save(output_xlsx_path)
            logger.user("所有任务处理完成！", 'SUCCESS')

            return render_template('index.html', 
                                   success=True, 
                                   logs=logger.logs,
                                   alert_csv='alert_result.csv', 
                                   swapped_csv='swapped_log.csv',
                                   result_xlsx='result_plan.xlsx',
                                   alert_count=len(alert_groups),
                                   secondary_alert_count=len(secondary_alert_groups),
                                   swap_count=actual_swapped_count,
                                   alert_preview=alert_groups, # Pass all for preview
                                   secondary_alert_preview=secondary_alert_groups,
                                   swap_preview=swapped_log_data) # Pass all for preview

        except Exception as e:
            traceback.print_exc()
            return f"Error: {str(e)}", 500

    return render_template('index.html')

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(app.config['DOWNLOAD_FOLDER'], filename, as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)
