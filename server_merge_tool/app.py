import os
import sys
import pandas as pd
import numpy as np
import datetime
from flask import Flask, render_template, request, send_file, send_from_directory
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import traceback

app = Flask(__name__)

# Determine if running as a script or frozen (PyInstaller)
if getattr(sys, 'frozen', False):
    BASE_DIR = sys._MEIPASS
    # When frozen, templates and static files are in sys._MEIPASS
    app = Flask(__name__, template_folder=os.path.join(BASE_DIR, 'templates'))
else:
    BASE_DIR = os.path.abspath(os.path.dirname(__file__))
    app = Flask(__name__)

UPLOAD_FOLDER = os.path.join(os.getcwd(), 'uploads') # Use CWD for user-accessible folders
DOWNLOAD_FOLDER = os.path.join(os.getcwd(), 'downloads')
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(DOWNLOAD_FOLDER, exist_ok=True)

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['DOWNLOAD_FOLDER'] = DOWNLOAD_FOLDER

class ExecutionLogger:
    def __init__(self):
        self.logs = []
    
    def user(self, message, level='INFO'):
        self._add_log(level, message, 'user')
        
    def dev(self, message, level='DEBUG'):
        self._add_log(level, message, 'dev')
        
    def _add_log(self, level, message, category):
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        self.logs.append({
            'time': timestamp, 
            'level': level, 
            'msg': message,
            'category': category
        })

def parse_server_pairs(text):
    pairs = []
    seen = set()
    duplicates = []
    
    if not text:
        return pairs, duplicates
        
    lines = text.strip().split('\n')
    for line in lines:
        parts = line.replace('，', ',').split(',')
        if len(parts) >= 2:
            try:
                s1 = int(parts[0].strip())
                s2 = int(parts[1].strip())
                
                # Sort tuple to treat (A, B) same as (B, A)
                pair_key = tuple(sorted((s1, s2)))
                
                if pair_key in seen:
                    duplicates.append(f"{s1} ↔ {s2}")
                else:
                    seen.add(pair_key)
                    pairs.append((s1, s2))
            except ValueError:
                continue
    return pairs, duplicates

def get_server_info(df, server_id):
    row = df[df['区服ID'] == server_id]
    if row.empty:
        return None
    return row.iloc[0]

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        logger = ExecutionLogger()
        try:
            logger.user("开始处理任务...")
            logger.dev("初始化请求参数解析")
            
            # 1. Save files
            csv_files = request.files.getlist('csv_files')
            xlsx_file = request.files['xlsx_file']
            pairs_text = request.form['pairs_text']

            if not csv_files or not xlsx_file:
                return "Missing files", 400

            xlsx_path = os.path.join(app.config['UPLOAD_FOLDER'], 'input.xlsx')
            xlsx_file.save(xlsx_path)
            logger.user("合服计划表 (XLSX) 上传成功")

            # 2. Process CSVs (Merge Multiple)
            logger.user(f"正在处理 {len(csv_files)} 个服务器数据文件...")
            dfs = []
            for i, file in enumerate(csv_files):
                if file.filename == '':
                    continue
                temp_path = os.path.join(app.config['UPLOAD_FOLDER'], f'input_{i}.csv')
                file.save(temp_path)
                try:
                    # 尝试读取 CSV，跳过第一行
                    df_temp = pd.read_csv(temp_path, header=1)
                    dfs.append(df_temp)
                    logger.dev(f"读取 CSV {file.filename} 成功，行数: {len(df_temp)}")
                except Exception as e:
                    logger.user(f"读取文件 {file.filename} 失败", 'ERROR')
                    logger.dev(f"CSV 读取异常: {str(e)}", 'ERROR')
            
            if not dfs:
                 return "没有有效的 CSV 文件", 400
                 
            df = pd.concat(dfs, ignore_index=True)
            logger.user(f"数据合并完成，共 {len(df)} 条记录")
            
            # Ensure numeric columns
            cols_to_numeric = ['区服ID', '前2名战力之和', '最高玩家累充金额', 'DAU']
            for col in cols_to_numeric:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            
            # Sort
            logger.dev("执行数据排序: 前2名战力之和 (降序)")
            df = df.sort_values(by='前2名战力之和', ascending=False).reset_index(drop=True)
            df['真实排名'] = df.index + 1
            total_servers = len(df)
            
            input_pairs, duplicates = parse_server_pairs(pairs_text)
            
            if duplicates:
                logger.user(f"发现并忽略 {len(duplicates)} 组重复检测对", 'WARN')
                if len(duplicates) <= 5:
                    for dup in duplicates:
                        logger.dev(f"忽略重复: {dup}", 'WARN')
                else:
                    logger.dev(f"重复列表 (前5个): {', '.join(duplicates[:5])}...", 'WARN')
            
            logger.user(f"解析输入：共 {len(input_pairs)} 组有效检测区服")
            
            alert_groups = [] 
            normal_groups = [] 
            
            # 3. Primary Alert Check
            logger.dev("开始执行初级警报检测 (Primary Check)")
            for s1, s2 in input_pairs:
                row1 = get_server_info(df, s1)
                row2 = get_server_info(df, s2)
                
                if row1 is None:
                    logger.user(f"警告：区服 {s1} 数据缺失，已跳过", 'WARN')
                if row2 is None:
                    logger.user(f"警告：区服 {s2} 数据缺失，已跳过", 'WARN')

                if row1 is None or row2 is None:
                    continue
                    
                rank1 = row1['真实排名']
                rank2 = row2['真实排名']
                
                # Conditions
                cond_a = abs(rank1 - rank2) <= 5
                
                top_25_threshold = total_servers * 0.25
                cond_b = (rank1 <= top_25_threshold and rank2 <= top_25_threshold and
                          row1['最高玩家累充金额'] >= 5000 and row2['最高玩家累充金额'] >= 5000)
                          
                power1 = row1['前2名战力之和']
                power2 = row2['前2名战力之和']
                cond_c = abs(power1 - power2) <= 1000000000
                
                if cond_a or cond_b or cond_c:
                    reasons = []
                    if cond_a: reasons.append(f"排名接近(差{abs(rank1-rank2)})")
                    if cond_b: reasons.append("高战高充(前25%)")
                    if cond_c: reasons.append("战力接近(差<=10亿)")
                    reason_str = "; ".join(reasons)
                    
                    logger.user(f"发现警报：{s1} 和 {s2} - {reason_str}", 'WARN')
                    alert_groups.append({'ids': [s1, s2], 'reason': reason_str})
                else:
                    normal_groups.append((s1, s2))
            
            logger.user(f"检测完成：发现 {len(alert_groups)} 组警报，{len(normal_groups)} 组正常")

            # 4. Secondary Alert Check
            logger.dev("加载 XLSX 进行二次关联检测")
            wb = load_workbook(xlsx_path)
            ws = wb.active
            
            header_row = [cell.value for cell in ws[1]]
            try:
                target_col_idx = header_row.index('目标服') 
                part_col_idx = header_row.index('参与服')
            except ValueError:
                target_col_idx = 0
                part_col_idx = 1

            def find_partner_in_xlsx(server_id):
                for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    t_id = row[target_col_idx]
                    p_id = row[part_col_idx]
                    if t_id == server_id: return r_idx, p_id
                    if p_id == server_id: return r_idx, t_id
                return None, None

            final_alert_rows = [] 
            secondary_alert_groups = []
            
            for group in alert_groups:
                group_id = f"Group_{group['ids'][0]}_{group['ids'][1]}"
                for sid in group['ids']:
                    r = get_server_info(df, sid)
                    if r is not None:
                        r_dict = r.to_dict()
                        r_dict['警报组ID'] = group_id
                        r_dict['警报原因'] = group['reason']
                        final_alert_rows.append(r_dict)
                        
                s1, s2 = group['ids'][0], group['ids'][1]
                partners = []
                _, p1 = find_partner_in_xlsx(s1)
                if p1: partners.append(p1)
                _, p2 = find_partner_in_xlsx(s2)
                if p2: partners.append(p2)
                
                dau_alerts = []

                # Check S1
                r1 = get_server_info(df, s1)
                if r1 is not None and r1['DAU'] <= 5:
                    dau_alerts.append(f"{s1}本身(DAU:{int(r1['DAU'])})")
                
                # Check S2
                r2 = get_server_info(df, s2)
                if r2 is not None and r2['DAU'] <= 5:
                    dau_alerts.append(f"{s2}本身(DAU:{int(r2['DAU'])})")

                # Check S1 Partner
                _, p1 = find_partner_in_xlsx(s1)
                if p1: 
                    partners.append(p1)
                    rp1 = get_server_info(df, p1)
                    if rp1 is not None and rp1['DAU'] <= 5:
                         dau_alerts.append(f"{s1}关联服{p1}(DAU:{int(rp1['DAU'])})")

                # Check S2 Partner
                _, p2 = find_partner_in_xlsx(s2)
                if p2: 
                    partners.append(p2)
                    rp2 = get_server_info(df, p2)
                    if rp2 is not None and rp2['DAU'] <= 5:
                         dau_alerts.append(f"{s2}关联服{p2}(DAU:{int(rp2['DAU'])})")
                
                if dau_alerts:
                    details = ", ".join(dau_alerts)
                    logger.user(f"组 {group_id} 触发二次警报: DAU过低 [{details}]", 'WARN')
                    logger.dev(f"组 {group_id} 触发二次警报 (DAU<=5) - {details}")
                    
                    secondary_alert_groups.append({
                        'ids': [s1, s2],
                        'reason': f"关联服DAU过低: {details}"
                    })

                    for pid in partners:
                        pr = get_server_info(df, pid)
                        if pr is not None:
                            pr_dict = pr.to_dict()
                            pr_dict['警报组ID'] = group_id
                            pr_dict['警报原因'] = f"二次查询DAU过低 [{details}]"
                            final_alert_rows.append(pr_dict)

            # Create Alert CSV with optimized formatting
            if final_alert_rows:
                alert_df = pd.DataFrame(final_alert_rows)
                
                # Reorder columns: Put '真实排名' first, then '警报组ID', '警报原因'
                cols = alert_df.columns.tolist()
                # Ensure '真实排名' is in columns (it was added during processing)
                priority_cols = ['真实排名', '警报组ID', '警报原因']
                # Filter out priority cols from existing cols to avoid duplication/error if missing
                priority_cols = [c for c in priority_cols if c in cols]
                other_cols = [c for c in cols if c not in priority_cols]
                alert_df = alert_df[priority_cols + other_cols]
                
                # Add empty rows between groups for visual separation
                # Convert to list of dicts to easily insert rows
                # Sort by Group ID to ensure they are contiguous
                alert_df.sort_values(by='警报组ID', inplace=True)
                
                output_rows = []
                current_group = None
                
                for _, row in alert_df.iterrows():
                    if current_group is not None and row['警报组ID'] != current_group:
                        # Insert empty row (dict with all NaN/None)
                        empty_row = {c: None for c in alert_df.columns}
                        output_rows.append(empty_row)
                    
                    output_rows.append(row.to_dict())
                    current_group = row['警报组ID']
                
                final_df = pd.DataFrame(output_rows)
                output_csv_path = os.path.join(app.config['DOWNLOAD_FOLDER'], 'alert_result.csv')
                final_df.to_csv(output_csv_path, index=False, encoding='utf-8-sig')
            else:
                pd.DataFrame().to_csv(os.path.join(app.config['DOWNLOAD_FOLDER'], 'alert_result.csv'), index=False)

            # 5. Swap Servers
            logger.user("正在处理正常组的交换逻辑...")
            fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            swapped_log_data = [] 

            server_row_map = {}
            for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                t_id = row[target_col_idx]
                p_id = row[part_col_idx]
                if t_id: server_row_map[t_id] = r_idx
                if p_id: server_row_map[p_id] = r_idx
                
            actual_swapped_count = 0
            
            for s1, s2 in normal_groups:
                r1_idx = server_row_map.get(s1)
                r2_idx = server_row_map.get(s2)
                
                if r1_idx and r2_idx and r1_idx != r2_idx:
                    actual_swapped_count += 1
                    
                    # Capture State Before Swap
                    c1_t = ws.cell(row=r1_idx, column=target_col_idx+1)
                    c1_p = ws.cell(row=r1_idx, column=part_col_idx+1)
                    v1_t, v1_p = c1_t.value, c1_p.value
                    
                    c2_t = ws.cell(row=r2_idx, column=target_col_idx+1)
                    c2_p = ws.cell(row=r2_idx, column=part_col_idx+1)
                    v2_t, v2_p = c2_t.value, c2_p.value
                    
                    # Helper to format server pair string
                    def fmt_pair(t, p):
                        return f"[{t if t else '空'} + {p if p else '空'}]"

                    before_str_1 = fmt_pair(v1_t, v1_p)
                    before_str_2 = fmt_pair(v2_t, v2_p)

                    # Swap Logic
                    new_v1_t = s2 if v1_t == s1 else v1_t
                    new_v1_p = s2 if v1_p == s1 else v1_p
                    new_v2_t = s1 if v2_t == s2 else v2_t
                    new_v2_p = s1 if v2_p == s2 else v2_p
                    
                    pair1 = sorted([x for x in [new_v1_t, new_v1_p] if x is not None])
                    if len(pair1) == 2:
                        c1_t.value, c1_p.value = pair1[0], pair1[1]
                        final_v1_t, final_v1_p = pair1[0], pair1[1]
                    else:
                        # Fallback if something is weird, though logic above tries to keep structure
                        c1_t.value, c1_p.value = new_v1_t, new_v1_p
                        final_v1_t, final_v1_p = new_v1_t, new_v1_p
                    
                    pair2 = sorted([x for x in [new_v2_t, new_v2_p] if x is not None])
                    if len(pair2) == 2:
                        c2_t.value, c2_p.value = pair2[0], pair2[1]
                        final_v2_t, final_v2_p = pair2[0], pair2[1]
                    else:
                         c2_t.value, c2_p.value = new_v2_t, new_v2_p
                         final_v2_t, final_v2_p = new_v2_t, new_v2_p
                        
                    # Capture State After Swap
                    after_str_1 = fmt_pair(final_v1_t, final_v1_p)
                    after_str_2 = fmt_pair(final_v2_t, final_v2_p)
                    
                    # Human readable change log
                    change_log = (
                        f"组1 (行{r1_idx}): {before_str_1} ➔ {after_str_1}\n"
                        f"   组2 (行{r2_idx}): {before_str_2} ➔ {after_str_2}"
                    )
                    
                    logger.user(f"✅ 成功交换 {s1} ↔ {s2}\n   {change_log}", 'SUCCESS')
                    
                    swapped_log_data.append({
                        '交换区服1': s1, '交换区服2': s2,
                        '原始行号1': r1_idx, '原始行号2': r2_idx,
                        'Before1': before_str_1, 'After1': after_str_1,
                        'Before2': before_str_2, 'After2': after_str_2,
                        '状态': '已交换'
                    })
                        
                    for cell in ws[r1_idx]: cell.fill = fill
                    for cell in ws[r2_idx]: cell.fill = fill

                    server_row_map[s1] = r2_idx
                    server_row_map[s2] = r1_idx
                    logger.dev(f"执行交换 ({s1}, {s2}) - Rows: {r1_idx} <-> {r2_idx}")
                else:
                    logger.dev(f"无法交换 ({s1}, {s2}): 未找到匹配行")

            if swapped_log_data:
                swapped_df = pd.DataFrame(swapped_log_data)
                output_swapped_path = os.path.join(app.config['DOWNLOAD_FOLDER'], 'swapped_log.csv')
                swapped_df.to_csv(output_swapped_path, index=False, encoding='utf-8-sig')
            else:
                pd.DataFrame().to_csv(os.path.join(app.config['DOWNLOAD_FOLDER'], 'swapped_log.csv'), index=False)

            output_xlsx_path = os.path.join(app.config['DOWNLOAD_FOLDER'], 'result_plan.xlsx')
            wb.save(output_xlsx_path)
            logger.user("所有任务处理完成！", 'SUCCESS')

            return render_template('index.html', 
                                   success=True, 
                                   logs=logger.logs,
                                   alert_csv='alert_result.csv', 
                                   swapped_csv='swapped_log.csv',
                                   result_xlsx='result_plan.xlsx',
                                   alert_count=len(alert_groups),
                                   secondary_alert_count=len(secondary_alert_groups),
                                   swap_count=actual_swapped_count,
                                   alert_preview=alert_groups[:10], # Pass top 10 for preview
                                   secondary_alert_preview=secondary_alert_groups[:10],
                                   swap_preview=swapped_log_data[:10]) # Pass top 10 for preview

        except Exception as e:
            traceback.print_exc()
            return f"Error: {str(e)}", 500

    return render_template('index.html')

@app.route('/download/<filename>')
def download_file(filename):
    return send_from_directory(app.config['DOWNLOAD_FOLDER'], filename, as_attachment=True)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5001, debug=True)
